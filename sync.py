"""
Syncs Portico Content Calendar from SharePoint to data.json.
Requires env vars: AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET,
                   SHAREPOINT_DRIVE_ID, SHAREPOINT_ITEM_ID
"""
import io, json, os
from datetime import datetime, timezone
import requests, msal
from openpyxl import load_workbook

TENANT_ID    = os.environ['AZURE_TENANT_ID']
CLIENT_ID    = os.environ['AZURE_CLIENT_ID']
CLIENT_SECRET= os.environ['AZURE_CLIENT_SECRET']
DRIVE_ID     = os.environ['SHAREPOINT_DRIVE_ID']
ITEM_ID      = os.environ['SHAREPOINT_ITEM_ID']

DATA_FILE = 'data.json'

FASFAA_L = 'https://simplifyedinc.sharepoint.com/:w:/s/Marketing/IQDJy9q5tvFDTYAecn8zVLvtASMPasFobtFVhLA6Tk9ZEDk?e=H'
ACEN_L   = 'https://simplifyedinc.sharepoint.com/:w:/s/Marketing/IQCOwWZ57RExRpfoOdwRK5cDAZlryGrBp-J5UEjlmdE9CCU?e'
CECU_L   = 'https://simplifyedinc.sharepoint.com/:w:/s/Marketing/IQAOqTCKwQyUQIcJ9jTgdAFgAfvrCs_ql78gYyqJaiyzLB0?e=Fme'

KNOWN_LINKS = {
    'FASFAA 2026': FASFAA_L,
    'ACEN 2026':   ACEN_L,
    'CECU 2026':   CECU_L,
}


def get_token():
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=f'https://login.microsoftonline.com/{TENANT_ID}',
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(['https://graph.microsoft.com/.default'])
    if 'access_token' not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description')}")
    return result['access_token']


def download_workbook(token):
    url = f'https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content'
    r = requests.get(url, headers={'Authorization': f'Bearer {token}'})
    r.raise_for_status()
    return load_workbook(io.BytesIO(r.content), data_only=True)


def parse_date(val):
    if not val:
        return None
    s = str(val).strip()
    for fmt in ('%d-%b', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            d = datetime.strptime(s.split(' ')[0] if ' ' in s else s, fmt)
            if d.year == 1900:
                d = d.replace(year=2026)
            return d.strftime('2026-%m-%d') if fmt == '%d-%b' else d.strftime('%Y-%m-%d')
        except:
            pass
    return None


def normalize_type(title, raw):
    t = (raw or '').strip()
    if 'Pre-Event' in title or 'Post-Event' in title:
        return 'Event Comms'
    mapping = {
        'Landing Page':     'Website',
        'LinkedIn':         'Social',
        'Case Study / Guide':'Case Study',
        'One-Pager':        'Sales Asset',
        'Product':          'Product Update',
    }
    return mapping.get(t, t)


def load_existing():
    if not os.path.exists(DATA_FILE):
        return {}, []
    with open(DATA_FILE) as f:
        d = json.load(f)
    meta = {}
    for item in d.get('items', []):
        if item.get('type') == 'Event':
            meta[(item.get('date'), item['title'][:8])] = {
                'endDate':    item.get('endDate', ''),
                'attendees':  item.get('attendees', ''),
                'contentLink':item.get('contentLink', ''),
            }
    return d, meta


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    # Find header row (contains 'Content' or 'Tags')
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(c or '').strip() for c in row]
        if 'Content' in vals and 'Tags' in vals:
            header_row = i
            break
    if header_row is None:
        raise RuntimeError("Could not find header row")

    headers = [str(c or '').strip() for c in rows[header_row]]
    col = {h: i for i, h in enumerate(headers) if h}

    def g(row, name, fallback=''):
        i = col.get(name)
        return str(row[i] or '').strip() if i is not None and i < len(row) else fallback

    items = []
    for row in rows[header_row + 1:]:
        title = g(row, 'Content')
        if not title:
            continue
        audience   = g(row, 'Audience')
        tags       = g(row, 'Tags')
        raw_type   = g(row, 'Content Type/Channel')
        sender     = g(row, 'Brand / Sender')
        date_val   = g(row, 'Date')
        end_val    = g(row, 'End date (events)')
        link       = g(row, 'Link')
        attendees  = g(row, 'Attendees (Events only)')

        t = normalize_type(title, raw_type)
        if sender == 'Portico Product Updates' and t == 'Email':
            t = 'Product Comms'

        item = {
            'date':        parse_date(date_val),
            'tags':        tags,
            'title':       title,
            'type':        t,
            'audience':    audience,
            'sender':      sender,
            'summary':     '',
            'audienceLink':'',
            'contentLink': link,
        }

        if t == 'Event':
            end_date = parse_date(end_val)
            if end_date:
                item['endDate'] = end_date
            if attendees:
                item['attendees'] = attendees
            # Known event brief links
            for prefix, url in KNOWN_LINKS.items():
                if title.startswith(prefix):
                    item['contentLink'] = url
                    break

        items.append(item)

    return items


def merge_meta(items, meta):
    for item in items:
        if item.get('type') != 'Event':
            continue
        key = (item.get('date'), item['title'][:8])
        old = meta.get(key, {})
        if not item.get('endDate') and old.get('endDate'):
            item['endDate'] = old['endDate']
        if not item.get('attendees') and old.get('attendees'):
            item['attendees'] = old['attendees']
        if not item.get('contentLink') and old.get('contentLink'):
            item['contentLink'] = old['contentLink']
    return items


def main():
    old_data, meta = load_existing()
    token = get_token()
    wb    = download_workbook(token)
    ws    = wb.active
    items = parse_sheet(ws)
    items = merge_meta(items, meta)

    out = {
        'lastUpdated': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'campaigns':   old_data.get('campaigns', []),
        'items':       items,
    }
    with open(DATA_FILE, 'w') as f:
        json.dump(out, f, indent=2)
    print(f"Synced {len(items)} items.")


if __name__ == '__main__':
    main()
